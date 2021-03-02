# -*- coding: utf-8 -*-
"""
Created on Sun Feb 28 20:54:17 2021

Noyau - Module 1

Exceptionnellement sans classe

@author: JCV
"""
#%% Récupération chemin actuel du script
def ouSuisJe():
    """
    cette fonction retourne le chemin complet du script en cours d'exécution'

    Returns
    -------
    chemin : str
        Chemin complet du script en cours d'exécution.
        Exemple : c:/user/jcv/memolab/

    """
    import os
    absFilePath = os.path.abspath(__file__)
    chemin, nomScript = os.path.split(absFilePath)
    chemin = chemin.replace("\\","/")
    return chemin


def chargeParametresGeneraux(nomFichierParametres):
    """
    Charge les paramètres contenus dans le fichier xlsx appelé dans le
    dictionnaire paraGen retourné au script appelant

    Parameters
    ----------
    nomFichierParametres : str
        Nom du fichier excel xlsx contenant la liste des paramètres selon
        le modèle fourni avec l'appli. Ex: params_generaux.xlsx

    Returns
    -------
    paraGen : dict
        Un terme par paramètre.
        L'appel d'un paramètre se fait ainsi :
            <valeurParamètre> = paraGen["<nomParamètre>"]

    """
    paraGen = {}
    # Accès au fichier excel
    import openpyxl
    # chemin du script courant
    chemin = ouSuisJe()
    # print(chemin)
    # objet workbook
    wb = openpyxl.load_workbook(chemin + "/" + nomFichierParametres)
    # objet worksheet
    ws = wb.active
    for ligne in range(1, 100):
        adrCelluleNom = "B" + str(ligne)
        adrCelluleVal = "C" + str(ligne)
        nomParametre = ws[adrCelluleNom].value
        valeurParametre = ws[adrCelluleVal].value
        if valeurParametre == None:
            break
        # ajout parametre au dictionnaire
        paraGen.update({nomParametre : valeurParametre})
        # essai création variable
        # exec(nomParametre + " = valeurParametre")
    wb.close()
    # print("JeanClaude =", JeanClaude)
    return paraGen

def xlsx2List(nomFichierXlsx,listXlsx,paraGen):
    """ Transfert du contenu de la feuille xlsx vers la liste 
    
    Paramètres:
        nomFichierXlsx : str : exemple: user.xlsx
        listXlsx : list
        paraGen : Dictionnaire des paramètres généraux
    
    Le nom complet avec chemin est formé ici au moyen de la
    fonction OuSuisJe() et du paraGen du nom du sous-pépertoire
    relatif contenant le fichier à ouvrir.
    
    Le nom de liste passé en paramètre doit correspondre à une
    liste existante, dont le contenu sera substitué par celui
    du fichier excel, ou alors initialisé au préalable par 
    l'appelant.
    
    Cette fonction retourne la liste listXlsx
    """
    
    # Etablissement du nom de fichier xlsx complet
    racineChemin = ouSuisJe()
    racineCheminSRep = racineChemin + paraGen["cheminData"] + "/"
    nomXlsxComplet = racineCheminSRep + nomFichierXlsx
    # if paraGen["imprimeOK"]:
    #     print("xlsx2List : nomXlsxComplet : {}".format(nomXlsxComplet))
    
    # accès au fichier excel et à sa feuille active
    import openpyxl
    wb = openpyxl.load_workbook(nomXlsxComplet,data_only=True)
    ws = wb.active
    
    # chargement de la feuille dans la liste
    for i, row in enumerate(ws.rows):
            listXlsx.append([])
            for cell in row:
                listXlsx[i].append(cell.value)
    
    # cloture accès fichier xlsx
    wb.close()

    return listXlsx    
    
""" # ################# code test xlsx2List #############
# charge paraGen
nomFichierParametres = "params_generaux.xlsx"
paraGen = chargeParametresGeneraux(nomFichierParametres)
nomFichierXlsx = "classeurExcel1.xlsx"
listXlsx=[]
listXlsx = xlsx2List(nomFichierXlsx,listXlsx,paraGen)
if paraGen["imprimeOK"]:
    print(listXlsx)
""" # ###################################################    
    

def list2Xlsx(listXlsx,nomFichierXlsx,paraGen):
    """ Transfert du contenu d'une liste xlsx vers un fichier Excel 
    
    Paramètres:
        listXlsx : list
        nomFichierXlsx : str : exemple: user.xlsx
        paraGen : Dictionnaire des paramètres généraux
    
    Le nom complet avec chemin est formé ici au moyen de la
    fonction OuSuisJe() et du paraGen du nom du sous-pépertoire
    relatif contenant le fichier à ouvrir.
    
    Le nom de liste passé en paramètre doit correspondre à une
    liste existante. Sa structure sera compatible avec celle de la
    feuille excel destinataire. Il s'agira typiquement d'une feuille
    excel qui a été chargée dans la liste, puis cette liste a
    été modifiée par le script. Et on actualise ici ces changements
    sur la feuille excel d'origine.
    
    Cette fonction ne retourne rien
    """    
    # Etablissement du nom de fichier xlsx complet
    racineChemin = ouSuisJe()
    racineCheminSRep = racineChemin + paraGen["cheminData"] + "/"
    nomXlsxComplet = racineCheminSRep + nomFichierXlsx
    # if paraGen["imprimeOK"]:
    #     print("xlsx2List : nomXlsxComplet : {}".format(nomXlsxComplet))
        
    # accès au fichier excel et à sa feuille active
    import openpyxl
    wb = openpyxl.load_workbook(nomXlsxComplet,data_only=True)
    ws = wb.active
    
    # chargement de la liste dans la feuille

    """ Evaluation nbre de lignes et de colonnes à transférer
    Partons du principe que nous traitons une liste 2D rectangulaire
    dont toutes les cellules de la ligne 0 contiennent quelque chose.
    Il en résultera que le nombre de lignes de la liste sera obtenue
    par la fonction len(list) et que le nombre de colonnes de la liste
    sera obtenu par la fonction len(list[0])
    """
    nbreLignes = len(listXlsx)
    nbreColonnes = len(listXlsx[0])

    # boucle écriture
    for ligne in range(1,nbreLignes):
        for colonne in range(1,nbreColonnes):
            ws.cell(row=ligne, column=colonne, value=listXlsx[ligne-1][colonne-1])
    
    # Sauvegarde de la feuille excel modifiée
    wb.save(nomXlsxComplet)
    
 
"""# ##### code test xlsx2List et list2Xlsx #############
# charge paraGen
nomFichierParametres = "params_generaux.xlsx"
paraGen = chargeParametresGeneraux(nomFichierParametres)
nomFichierXlsx = "essai.xlsx"
listXlsx=[]
# charge liste 
listXlsx = xlsx2List(nomFichierXlsx,listXlsx,paraGen)
if paraGen["imprimeOK"]:
    print(listXlsx)
    
# modification d'un élément de la liste
listXlsx[0][0] = "Origine"

# sauve liste dans fichier Excel
list2Xlsx(listXlsx,nomFichierXlsx,paraGen)
"""# ###################################################  

    
    
    
    
    
    
    
    